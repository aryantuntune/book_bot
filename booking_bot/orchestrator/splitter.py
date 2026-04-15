"""Turn a big xlsx into N smaller xlsx chunks, preserving the header row
and row order. Output lives under Input/chunks/<source>/<chunk-id>.xlsx.
Pure openpyxl — no orchestrator state, no subprocess."""
from __future__ import annotations

import logging
import re
import sys
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from booking_bot import config

log = logging.getLogger("orchestrator.splitter")

# Source name max length of 28 leaves headroom for "<source>-<NNN>" to
# stay under the existing bot CLI's 32-char --profile-suffix cap.
_SOURCE_RE = re.compile(r"^[A-Za-z0-9_-]{1,28}$")


@dataclass(frozen=True)
class ChunkSpec:
    source: str
    chunk_id: str
    chunk_index: int
    input_path: Path
    profile_suffix: str
    heartbeat_path: Path
    row_count: int
    operator_slot: str = "op1"
    operator_phone: str = ""


def _validate_source(source: str) -> None:
    if not _SOURCE_RE.fullmatch(source):
        raise ValueError(
            f"source must match {_SOURCE_RE.pattern!r} "
            f"(alphanumeric + dash + underscore, 1-28 chars); got {source!r}"
        )


def split(
    source: str,
    input_file: Path,
    *,
    chunk_size: int | None = None,
    num_chunks: int | None = None,
    operator_phones: list[str] | None = None,
    clones_per_operator: int = 3,
    output_dir: Path | None = None,
) -> list[ChunkSpec]:
    """Split input_file into chunks.

    Two modes:
      - Single-operator (legacy): pass exactly one of chunk_size /
        num_chunks. All chunks get operator_slot='op1', operator_phone=''.
      - Multi-operator: pass operator_phones=[...]. Produces
        K*clones_per_operator contiguous chunks. First M chunks get
        operator_slot='op1' and operator_phone=phones[0], next M -> op2,
        and so on. `chunk_size`/`num_chunks` are ignored in this mode.

    Writes chunks to output_dir/<source>/<chunk-id>.xlsx (default
    output_dir = config.CHUNKS_DIR). Idempotent: skips writing a chunk
    whose row count already matches what's on disk."""
    _validate_source(source)

    if operator_phones is not None:
        if not operator_phones:
            raise ValueError("operator_phones must be non-empty")
        if not (1 <= clones_per_operator <= 3):
            raise ValueError(
                f"clones_per_operator must be between 1 and 3 (per-account "
                f"session limit); got {clones_per_operator}"
            )
        n_chunks_override = len(operator_phones) * clones_per_operator
        # Ignore chunk_size/num_chunks in multi-operator mode.
        chunk_size = None
        num_chunks = n_chunks_override
    else:
        if (chunk_size is None) == (num_chunks is None):
            raise ValueError(
                "pass exactly one of chunk_size or num_chunks to split()"
            )
        if chunk_size is not None and chunk_size <= 0:
            raise ValueError(f"chunk_size must be positive; got {chunk_size}")
        if num_chunks is not None and num_chunks <= 0:
            raise ValueError(f"num_chunks must be positive; got {num_chunks}")

    input_file = Path(input_file)
    out_root = Path(output_dir) if output_dir is not None else config.CHUNKS_DIR
    chunks_dir = out_root / source
    chunks_dir.mkdir(parents=True, exist_ok=True)

    header, data_rows = _read_input_rows(input_file)
    total_rows = len(data_rows)
    if total_rows == 0:
        raise ValueError(f"input file has no data rows: {input_file}")

    effective_size, n_chunks = _resolve_parallelism(
        total_rows, chunk_size=chunk_size, num_chunks=num_chunks,
    )
    pad_width = max(3, len(str(n_chunks)))
    if n_chunks > 50:
        print(f"[splitter] WARNING: num_chunks={n_chunks} is unusually high",
              file=sys.stderr)
    if effective_size < 10:
        print(f"[splitter] WARNING: chunk size={effective_size} is unusually low",
              file=sys.stderr)

    specs: list[ChunkSpec] = []
    for i in range(n_chunks):
        start = i * effective_size
        end = min(start + effective_size, total_rows)
        rows_slice = data_rows[start:end]
        chunk_index = i + 1
        chunk_id = f"{source}-{chunk_index:0{pad_width}d}"
        chunk_path = chunks_dir / f"{chunk_id}.xlsx"
        heartbeat_path = config.RUNS_DIR / source / f"{chunk_id}.heartbeat.json"
        _write_chunk_file(chunk_path, header, rows_slice)

        if operator_phones is not None:
            op_idx = i // clones_per_operator
            operator_slot = f"op{op_idx + 1}"
            operator_phone = operator_phones[op_idx]
        else:
            operator_slot = "op1"
            operator_phone = ""

        specs.append(ChunkSpec(
            source=source,
            chunk_id=chunk_id,
            chunk_index=chunk_index,
            input_path=chunk_path,
            profile_suffix=chunk_id,
            heartbeat_path=heartbeat_path,
            row_count=len(rows_slice),
            operator_slot=operator_slot,
            operator_phone=operator_phone,
        ))
    return specs


def _read_input_rows(path: Path) -> tuple[list, list[list]]:
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = list(rows[0])
    data = [list(r) for r in rows[1:] if any(v is not None for v in r)]
    wb.close()
    return header, data


def _resolve_parallelism(
    total_rows: int, *, chunk_size: int | None, num_chunks: int | None,
) -> tuple[int, int]:
    if chunk_size is not None:
        effective_size = chunk_size
        n = (total_rows + effective_size - 1) // effective_size
        return effective_size, n
    assert num_chunks is not None
    if num_chunks > total_rows:
        raise ValueError(
            f"num_chunks {num_chunks} exceeds row count {total_rows}"
        )
    effective_size = (total_rows + num_chunks - 1) // num_chunks
    return effective_size, num_chunks


def _write_chunk_file(path: Path, header: list, rows: list[list]) -> None:
    """Idempotent: if the file exists with the same data row count, skip."""
    if path.exists():
        try:
            wb_check = openpyxl.load_workbook(path, read_only=True)
            ws_check = wb_check.active
            existing_count = sum(
                1 for r in ws_check.iter_rows(min_row=2, values_only=True)
                if any(v is not None for v in r)
            )
            wb_check.close()
            if existing_count == len(rows):
                return
        except Exception:
            pass  # corrupt or unreadable — fall through and overwrite
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(path)
