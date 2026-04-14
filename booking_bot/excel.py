"""ExcelStore — the only module in the package that touches openpyxl / xlrd.

Design notes (see spec §7):
- Output is the source of truth for resume. Pending = col C empty, col B not None.
- pending_rows() does NOT validate; it yields raw cell values. The cli layer
  calls normalize_phone() on each.
- All writes go through atomic save: write to <path>.tmp then os.replace().
"""
from __future__ import annotations

import logging
import os
import shutil
from pathlib import Path
from typing import Iterator

import openpyxl

from booking_bot import config

log = logging.getLogger("excel")


class ExcelStore:
    def __init__(self, input_path: Path) -> None:
        self.input_path = Path(input_path)
        self.output_path = config.OUTPUT_DIR / self.input_path.name
        self.issues_path = config.ISSUES_DIR / self.input_path.name

        config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        config.ISSUES_DIR.mkdir(parents=True, exist_ok=True)

        # Handle .xls legacy input (§7.1). Normalize self.input_path to .xlsx.
        if self.input_path.suffix.lower() == ".xls":
            self.input_path = self._convert_xls_to_xlsx(self.input_path)
            self.output_path = config.OUTPUT_DIR / self.input_path.name

        if not self.output_path.exists():
            shutil.copy2(self.input_path, self.output_path)
            log.info(f"created Output workbook: {self.output_path}")
        else:
            log.info(f"resuming existing Output workbook: {self.output_path}")

        self._wb = openpyxl.load_workbook(self.output_path)
        self._ws = self._wb.active
        self._issues_wb: openpyxl.Workbook | None = None
        self._issues_ws = None

    def _convert_xls_to_xlsx(self, xls_path: Path) -> Path:
        """Read legacy .xls via xlrd and write out a .xlsx copy next to it."""
        import xlrd                              # lazy import — heavy
        book = xlrd.open_workbook(str(xls_path))
        sheet = book.sheet_by_index(0)
        wb = openpyxl.Workbook()
        ws = wb.active
        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                ws.cell(row=r + 1, column=c + 1).value = sheet.cell_value(r, c)
        new_path = xls_path.with_suffix(".xlsx")
        wb.save(new_path)
        log.info(f"converted {xls_path.name} -> {new_path.name}")
        return new_path

    # ---- Resume iteration ----

    def pending_rows(self) -> Iterator[tuple[int, object]]:
        """Yield (row_idx, raw_col_B_value) for rows where col B is not None AND
        col C is empty/whitespace. Operates on the Output workbook, starting at
        min_row=1 (no header row)."""
        for row in self._ws.iter_rows(min_row=1, values_only=False):
            row_idx = row[0].row
            col_b = row[1].value if len(row) > 1 else None
            col_c = row[2].value if len(row) > 2 else None
            if col_b is None:
                continue
            if col_c is not None and str(col_c).strip() != "":
                continue
            yield (row_idx, col_b)

    # ---- Writes ----

    def write_success(self, row_idx: int, code: str) -> None:
        """Write the 6-digit code to col C of row_idx, then atomically save."""
        self._ws.cell(row=row_idx, column=3).value = code
        self._atomic_save(self._wb, self.output_path)
        log.info(f"row {row_idx}: success code={code}")

    @staticmethod
    def _atomic_save(wb: openpyxl.Workbook, path: Path) -> None:
        """Save to <path>.tmp then os.replace — atomic on NTFS for same-FS
        renames. Never leaves a half-written .xlsx."""
        tmp = path.with_suffix(path.suffix + ".tmp")
        wb.save(tmp)
        os.replace(tmp, path)

    def write_issue(self, row_idx: int, phone: str, reason: str, raw: str) -> None:
        """Mark col C = 'ISSUE' in Output; append a row to Issues (create the
        workbook on first call). Issues columns:
          A: consumer number (mirrored from Output col A)
          B: phone (cleaned or raw cell, whatever the caller has)
          C: reason + raw chatbot text (joined with ' | ')
          D: cross-reference to Output row ('row N in Output/<file>')
        """
        self._ws.cell(row=row_idx, column=3).value = "ISSUE"
        self._atomic_save(self._wb, self.output_path)

        self._ensure_issues_workbook()
        consumer_no = self._ws.cell(row=row_idx, column=1).value
        # Find the next empty row in the Issues worksheet.
        next_row = 1
        while self._issues_ws.cell(row=next_row, column=1).value is not None:
            next_row += 1
        self._issues_ws.cell(row=next_row, column=1).value = consumer_no
        self._issues_ws.cell(row=next_row, column=2).value = phone
        self._issues_ws.cell(row=next_row, column=3).value = f"{reason} | {raw}"
        self._issues_ws.cell(row=next_row, column=4).value = (
            f"row {row_idx} in Output/{self.output_path.name}"
        )
        self._atomic_save(self._issues_wb, self.issues_path)
        log.warning(f"row {row_idx}: ISSUE ({reason})")

    def mark_terminal(self, row_idx: int, col_c_text: str) -> None:
        """Write a human-readable terminal label to col C (e.g. 'ekyc not
        done', 'booked'). Unlike write_issue, this does NOT touch the Issues
        workbook — the label itself tells the operator what happened and the
        row does not need a diagnostic dump. The outer retry loop treats
        non-numeric col C values as terminal, so these rows aren't retried."""
        self._ws.cell(row=row_idx, column=3).value = col_c_text
        self._atomic_save(self._wb, self.output_path)
        log.info(f"row {row_idx}: marked terminal = {col_c_text!r}")

    def clear_issue(self, row_idx: int) -> None:
        """Wipe col C for `row_idx` so pending_rows() picks it up again on
        the next pass. Used by the transient-retry loop in cli.py — a row
        that failed with a recoverable reason (unknown_state, playbook_stuck,
        etc.) gets its ISSUE marker cleared and is re-attempted. Terminal
        reasons (pending_payment, invalid_customer, already_booked) are left
        alone."""
        self._ws.cell(row=row_idx, column=3).value = None
        self._atomic_save(self._wb, self.output_path)
        log.info(f"row {row_idx}: cleared for retry")

    def _ensure_issues_workbook(self) -> None:
        """Lazily create or open the Issues workbook."""
        if self._issues_wb is not None:
            return
        if self.issues_path.exists():
            self._issues_wb = openpyxl.load_workbook(self.issues_path)
        else:
            self._issues_wb = openpyxl.Workbook()
        self._issues_ws = self._issues_wb.active

    def summary(self) -> dict[str, int]:
        """Classify every row by col C value so the CLI can log progress.

        Buckets:
          success        — col C is a pure 6-digit delivery confirmation code
          ekyc           — 'ekyc not done' terminal label
          not_registered — 'not registered with HPCL' terminal label
          payment_pending — 'payment pending' terminal label
          issue          — any other non-empty value (usually literal 'ISSUE',
                           but also future terminal labels we haven't special-
                           cased yet)
          pending        — col C is None or blank

        Returns a dict with all buckets plus `total`. `done` is a convenience
        alias for success + all terminal buckets — it's the progress bar
        numerator that stays stable across retry passes."""
        total = success = ekyc = not_registered = payment_pending = issue = pending = 0
        for row in self._ws.iter_rows(min_row=1, values_only=True):
            phone = row[1] if len(row) > 1 else None
            code  = row[2] if len(row) > 2 else None
            if phone is None:
                continue
            total += 1
            if code is None or str(code).strip() == "":
                pending += 1
                continue
            c = str(code).strip()
            if c.isdigit():
                success += 1
            elif c.lower() == "ekyc not done":
                ekyc += 1
            elif c.lower() == "not registered with hpcl":
                not_registered += 1
            elif c.lower() == "payment pending":
                payment_pending += 1
            else:
                issue += 1
        done = success + ekyc + not_registered + payment_pending + issue
        return {
            "total": total,
            "done": done,
            "success": success,
            "ekyc": ekyc,
            "not_registered": not_registered,
            "payment_pending": payment_pending,
            "issue": issue,
            "pending": pending,
        }

    def progress_line(self) -> str:
        """One-line human-readable progress for log output. Shows how many
        rows are resolved (done/total) with a per-bucket breakdown so the
        operator can see at a glance whether the remaining pending rows are
        shrinking. Example output:

            progress: 15/136 done (success=10 ekyc=2 not_reg=1 pay_pend=1 issue=1) pending=121
        """
        s = self.summary()
        return (
            f"progress: {s['done']}/{s['total']} done "
            f"(success={s['success']} ekyc={s['ekyc']} "
            f"not_reg={s['not_registered']} pay_pend={s['payment_pending']} "
            f"issue={s['issue']}) pending={s['pending']}"
        )
