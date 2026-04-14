"""TDD for ExcelStore.__init__ (copy input→output on first run, reuse existing
output on resume) and pending_rows (yields (row_idx, raw_cell) for rows with
empty col C and non-empty col B)."""
from pathlib import Path

import openpyxl
import pytest

from booking_bot.excel import ExcelStore


def _make_input(tmp_path: Path, rows: list[tuple]) -> Path:
    """Create an Input xlsx with columns A (consumer no.), B (phone)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(list(r))
    p = tmp_path / "Input" / "file1.xlsx"
    p.parent.mkdir(parents=True)
    wb.save(p)
    return p


@pytest.fixture()
def store_env(tmp_path, monkeypatch):
    """Point config paths at tmp_path and hand back the tmp_path."""
    from booking_bot import config
    monkeypatch.setattr(config, "INPUT_DIR",  tmp_path / "Input")
    monkeypatch.setattr(config, "OUTPUT_DIR", tmp_path / "Output")
    monkeypatch.setattr(config, "ISSUES_DIR", tmp_path / "Issues")
    (tmp_path / "Output").mkdir()
    (tmp_path / "Issues").mkdir()
    return tmp_path


def test_first_run_copies_input_to_output(store_env):
    inp = _make_input(store_env, [("C1", "9876543210"), ("C2", "9123456789")])
    store = ExcelStore(inp)
    out = store_env / "Output" / "file1.xlsx"
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "C1"
    assert ws.cell(row=1, column=2).value == "9876543210"


def test_pending_rows_yields_unfilled_rows(store_env):
    inp = _make_input(store_env, [
        ("C1", "9876543210"),
        ("C2", 9123456789),
        ("C3", ""),            # empty phone → skipped
        ("C4", "9000000000"),
    ])
    store = ExcelStore(inp)
    rows = list(store.pending_rows())
    assert [r[0] for r in rows] == [1, 2, 4]
    assert rows[0][1] == "9876543210"
    assert rows[1][1] == 9123456789
    assert rows[2][1] == "9000000000"


def test_pending_rows_skips_already_filled(store_env):
    """Rows where col C already has a value (code or 'ISSUE') are not yielded."""
    inp = _make_input(store_env, [
        ("C1", "9876543210"),
        ("C2", "9123456789"),
        ("C3", "9000000000"),
    ])
    # First run: create Output then manually mark row 2 as filled.
    store = ExcelStore(inp)
    out = store_env / "Output" / "file1.xlsx"
    wb = openpyxl.load_workbook(out)
    wb.active.cell(row=2, column=3).value = "764260"
    wb.save(out)

    store2 = ExcelStore(inp)
    pending = [r[0] for r in store2.pending_rows()]
    assert pending == [1, 3]


def test_pending_rows_skips_none_phone(store_env):
    inp = _make_input(store_env, [
        ("C1", None),
        ("C2", "9876543210"),
    ])
    store = ExcelStore(inp)
    assert [r[0] for r in store.pending_rows()] == [2]


def test_write_success_writes_code(store_env):
    inp = _make_input(store_env, [("C1", "9876543210"), ("C2", "9123456789")])
    store = ExcelStore(inp)

    store.write_success(1, "764260")

    wb = openpyxl.load_workbook(store_env / "Output" / "file1.xlsx")
    assert wb.active.cell(row=1, column=3).value == "764260"
    assert wb.active.cell(row=2, column=3).value in (None, "")


def test_write_success_is_atomic(store_env):
    """After a successful write, the .tmp sibling must not exist."""
    inp = _make_input(store_env, [("C1", "9876543210")])
    store = ExcelStore(inp)
    store.write_success(1, "111111")
    tmp = store_env / "Output" / "file1.xlsx.tmp"
    assert not tmp.exists()


def test_write_success_updates_pending_iteration(store_env):
    inp = _make_input(store_env, [
        ("C1", "9876543210"),
        ("C2", "9123456789"),
    ])
    store = ExcelStore(inp)
    store.write_success(1, "222222")

    store2 = ExcelStore(inp)
    assert [r[0] for r in store2.pending_rows()] == [2]


def test_write_issue_marks_output_and_creates_issues(store_env):
    inp = _make_input(store_env, [("C1", "9876543210")])
    store = ExcelStore(inp)

    store.write_issue(1, "9876543210", reason="unexpected_state",
                      raw="bot said: please try later")

    out = openpyxl.load_workbook(store_env / "Output" / "file1.xlsx")
    assert out.active.cell(row=1, column=3).value == "ISSUE"

    issues_path = store_env / "Issues" / "file1.xlsx"
    assert issues_path.exists()
    iss = openpyxl.load_workbook(issues_path)
    row = [iss.active.cell(row=1, column=c).value for c in range(1, 5)]
    assert row[0] == "C1"
    assert row[1] == "9876543210"
    assert "unexpected_state" in row[2]
    assert "please try later" in row[2]
    assert "row 1" in row[3]


def test_write_issue_appends_multiple(store_env):
    inp = _make_input(store_env, [("C1", "9876543210"), ("C2", "9123456789")])
    store = ExcelStore(inp)
    store.write_issue(1, "9876543210", reason="r1", raw="raw1")
    store.write_issue(2, "9123456789", reason="r2", raw="raw2")

    iss = openpyxl.load_workbook(store_env / "Issues" / "file1.xlsx")
    assert iss.active.max_row == 2
    assert iss.active.cell(row=2, column=2).value == "9123456789"


def test_summary_reports_counts(store_env):
    inp = _make_input(store_env, [
        ("C1", "9876543210"),
        ("C2", "9123456789"),
        ("C3", "9000000000"),
        ("C4", "9111111111"),
    ])
    store = ExcelStore(inp)
    store.write_success(1, "111111")
    store.write_issue(2, "9123456789", "unexpected_state", "raw")
    # rows 3, 4 still pending

    s = store.summary()
    assert s == {
        "total": 4,
        "done": 2,
        "success": 1,
        "ekyc": 0,
        "not_registered": 0,
        "payment_pending": 0,
        "issue": 1,
        "pending": 2,
    }


def test_summary_classifies_terminal_labels(store_env):
    inp = _make_input(store_env, [
        ("C1", "9876543210"),
        ("C2", "9123456789"),
        ("C3", "9000000000"),
        ("C4", "9111111111"),
        ("C5", "9222222222"),
        ("C6", "9333333333"),
        ("C7", "9444444444"),
        ("C8", "9555555555"),
    ])
    store = ExcelStore(inp)
    store.write_success(1, "719285")
    store.mark_terminal(2, "ekyc not done")
    store.mark_terminal(3, "not registered with HPCL")
    store.mark_terminal(4, "payment pending")
    store.write_issue(5, "9222222222", "unexpected_state", "raw")
    store.write_success(6, "719286")
    # rows 7, 8 still pending

    s = store.summary()
    assert s["total"] == 8
    assert s["success"] == 2
    assert s["ekyc"] == 1
    assert s["not_registered"] == 1
    assert s["payment_pending"] == 1
    assert s["issue"] == 1
    assert s["pending"] == 2
    assert s["done"] == 6
    # done should never double-count: success + terminals + issue
    assert s["done"] == s["success"] + s["ekyc"] + s["not_registered"] \
        + s["payment_pending"] + s["issue"]
    assert s["done"] + s["pending"] == s["total"]


def test_summary_terminal_label_matching_is_case_insensitive(store_env):
    inp = _make_input(store_env, [
        ("C1", "9876543210"),
        ("C2", "9123456789"),
        ("C3", "9000000000"),
    ])
    store = ExcelStore(inp)
    # Operator may hand-edit the file with mixed casing.
    import openpyxl
    wb = openpyxl.load_workbook(store.output_path)
    ws = wb.active
    ws.cell(row=1, column=3).value = "EKYC Not Done"
    ws.cell(row=2, column=3).value = "Not Registered With HPCL"
    ws.cell(row=3, column=3).value = "PAYMENT PENDING"
    wb.save(store.output_path)

    # Reload so the store sees the hand-edited labels.
    store2 = ExcelStore(inp)
    s = store2.summary()
    assert s["ekyc"] == 1
    assert s["not_registered"] == 1
    assert s["payment_pending"] == 1
    assert s["issue"] == 0


def test_progress_line_formats_all_buckets(store_env):
    inp = _make_input(store_env, [
        ("C1", "9876543210"),
        ("C2", "9123456789"),
        ("C3", "9000000000"),
    ])
    store = ExcelStore(inp)
    store.write_success(1, "719285")
    store.mark_terminal(2, "ekyc not done")
    # row 3 pending

    line = store.progress_line()
    # Must contain done/total, plus every bucket label so operators can
    # eyeball any category growing unexpectedly.
    assert "2/3 done" in line
    assert "success=1" in line
    assert "ekyc=1" in line
    assert "not_reg=0" in line
    assert "pay_pend=0" in line
    assert "issue=0" in line
    assert "pending=1" in line


def test_progress_line_on_empty_file(store_env):
    # Edge case: input file with zero data rows. progress_line should not
    # crash and should show 0/0 cleanly.
    inp = _make_input(store_env, [])
    store = ExcelStore(inp)
    s = store.summary()
    assert s["total"] == 0
    assert s["done"] == 0
    assert s["pending"] == 0
    line = store.progress_line()
    assert "0/0 done" in line
