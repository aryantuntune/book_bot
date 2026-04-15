"""Unit tests for orchestrator/splitter.py. Uses openpyxl to build
fixtures on the fly (no checked-in binaries)."""
from pathlib import Path

import openpyxl
import pytest

from booking_bot import config
from booking_bot.orchestrator import splitter


def _make_input_xlsx(path: Path, n_rows: int) -> Path:
    """Build a throwaway input workbook with 1 header row + n_rows of data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["consumer_no", "phone"])
    for i in range(n_rows):
        ws.append([f"C{i+1}", f"98765{i:05d}"])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


@pytest.fixture()
def split_env(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(config, "CHUNKS_DIR", tmp_path / "Input" / "chunks")
    monkeypatch.setattr(config, "RUNS_DIR", tmp_path / "data" / "runs")
    return tmp_path


def test_chunk_spec_fields_are_frozen():
    spec = splitter.ChunkSpec(
        source="TEST", chunk_id="TEST-001", chunk_index=1,
        input_path=Path("Input/chunks/TEST/TEST-001.xlsx"),
        profile_suffix="TEST-001",
        heartbeat_path=Path("data/runs/TEST/TEST-001.heartbeat.json"),
        row_count=5,
    )
    with pytest.raises(Exception):
        spec.source = "HACKED"  # type: ignore[misc]


@pytest.mark.parametrize("bad_source", [
    "", " ", "foo/bar", "has space", "foo.bar", "A" * 29,
])
def test_invalid_source_name_raises(split_env, bad_source):
    inp = _make_input_xlsx(split_env / "Input" / "file.xlsx", n_rows=5)
    with pytest.raises(ValueError, match="source"):
        splitter.split(bad_source, inp, chunk_size=5)


@pytest.mark.parametrize("good_source", [
    "ASU", "IOCL", "BPCL-feb", "indian_oil", "x", "A" * 28,
])
def test_valid_source_names_accepted(split_env, good_source):
    inp = _make_input_xlsx(split_env / "Input" / "file.xlsx", n_rows=5)
    chunks = splitter.split(good_source, inp, chunk_size=5)
    assert chunks[0].source == good_source


def test_split_by_chunk_size_creates_correct_chunk_count(split_env):
    inp = _make_input_xlsx(split_env / "Input" / "file.xlsx", n_rows=20)
    chunks = splitter.split("TEST", inp, chunk_size=5)
    assert len(chunks) == 4
    for i, c in enumerate(chunks, start=1):
        assert c.chunk_index == i
        assert c.row_count == 5
        assert c.chunk_id == f"TEST-{i:03d}"
        assert c.profile_suffix == f"TEST-{i:03d}"


def test_split_last_chunk_is_smaller(split_env):
    inp = _make_input_xlsx(split_env / "Input" / "file.xlsx", n_rows=23)
    chunks = splitter.split("TEST", inp, chunk_size=5)
    assert len(chunks) == 5
    assert [c.row_count for c in chunks] == [5, 5, 5, 5, 3]


def test_split_writes_chunk_files_with_header(split_env):
    inp = _make_input_xlsx(split_env / "Input" / "file.xlsx", n_rows=10)
    chunks = splitter.split("TEST", inp, chunk_size=5)
    for c in chunks:
        assert c.input_path.exists()
        wb = openpyxl.load_workbook(c.input_path)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "consumer_no"
        data_rows = sum(1 for row in ws.iter_rows(min_row=2)
                        if row[0].value is not None)
        assert data_rows == 5


def test_split_heartbeat_paths_are_under_runs_dir(split_env, monkeypatch):
    monkeypatch.setattr(config, "RUNS_DIR", split_env / "data" / "runs")
    inp = _make_input_xlsx(split_env / "Input" / "file.xlsx", n_rows=5)
    chunks = splitter.split("TEST", inp, chunk_size=5)
    assert chunks[0].heartbeat_path == (
        split_env / "data" / "runs" / "TEST" / "TEST-001.heartbeat.json"
    )


def test_split_refuses_when_neither_size_nor_num_chunks_given(split_env):
    inp = _make_input_xlsx(split_env / "Input" / "file.xlsx", n_rows=5)
    with pytest.raises(ValueError, match="exactly one"):
        splitter.split("TEST", inp)


def test_split_refuses_when_both_size_and_num_chunks_given(split_env):
    inp = _make_input_xlsx(split_env / "Input" / "file.xlsx", n_rows=5)
    with pytest.raises(ValueError, match="exactly one"):
        splitter.split("TEST", inp, chunk_size=5, num_chunks=2)


def test_split_refuses_zero_rows(split_env):
    inp = _make_input_xlsx(split_env / "Input" / "file.xlsx", n_rows=0)
    with pytest.raises(ValueError, match="no data rows"):
        splitter.split("TEST", inp, chunk_size=5)


@pytest.mark.parametrize("bad_size", [0, -1])
def test_split_refuses_nonpositive_chunk_size(split_env, bad_size):
    inp = _make_input_xlsx(split_env / "Input" / "file.xlsx", n_rows=5)
    with pytest.raises(ValueError):
        splitter.split("TEST", inp, chunk_size=bad_size)


def test_split_by_num_chunks_exact_division(split_env):
    inp = _make_input_xlsx(split_env / "Input" / "file.xlsx", n_rows=20)
    chunks = splitter.split("TEST", inp, num_chunks=4)
    assert len(chunks) == 4
    assert all(c.row_count == 5 for c in chunks)


def test_split_by_num_chunks_with_remainder(split_env):
    inp = _make_input_xlsx(split_env / "Input" / "file.xlsx", n_rows=23)
    chunks = splitter.split("TEST", inp, num_chunks=5)
    assert len(chunks) == 5
    # ceil(23/5)=5, so sizes are 5,5,5,5,3 (last absorbs the remainder).
    assert [c.row_count for c in chunks] == [5, 5, 5, 5, 3]


def test_split_num_chunks_exceeding_rows_raises(split_env):
    inp = _make_input_xlsx(split_env / "Input" / "file.xlsx", n_rows=5)
    with pytest.raises(ValueError, match="exceeds row count"):
        splitter.split("TEST", inp, num_chunks=10)


def test_split_single_chunk_holds_all_rows(split_env):
    inp = _make_input_xlsx(split_env / "Input" / "file.xlsx", n_rows=20)
    chunks = splitter.split("TEST", inp, num_chunks=1)
    assert len(chunks) == 1
    assert chunks[0].row_count == 20


def test_zero_padding_widens_for_many_chunks(split_env):
    inp = _make_input_xlsx(split_env / "Input" / "file.xlsx", n_rows=2400)
    chunks = splitter.split("TEST", inp, num_chunks=1200)
    assert chunks[0].chunk_id == "TEST-0001"
    assert chunks[-1].chunk_id == "TEST-1200"


def test_split_is_idempotent_on_rerun(split_env):
    inp = _make_input_xlsx(split_env / "Input" / "file.xlsx", n_rows=20)
    chunks1 = splitter.split("TEST", inp, chunk_size=5)
    mtimes1 = {c.input_path: c.input_path.stat().st_mtime_ns for c in chunks1}
    chunks2 = splitter.split("TEST", inp, chunk_size=5)
    mtimes2 = {c.input_path: c.input_path.stat().st_mtime_ns for c in chunks2}
    assert mtimes1 == mtimes2
    assert [c.chunk_id for c in chunks1] == [c.chunk_id for c in chunks2]


def test_chunk_spec_has_operator_slot_and_phone_defaults():
    spec = splitter.ChunkSpec(
        source="TEST", chunk_id="TEST-001", chunk_index=1,
        input_path=Path("Input/chunks/TEST/TEST-001.xlsx"),
        profile_suffix="TEST-001",
        heartbeat_path=Path("data/runs/TEST/TEST-001.heartbeat.json"),
        row_count=5,
    )
    assert spec.operator_slot == "op1"
    assert spec.operator_phone == ""


def test_split_multi_operator_K3_M3(split_env):
    inp = _make_input_xlsx(split_env / "Input" / "file.xlsx", n_rows=27)
    chunks = splitter.split(
        "LALJI", inp,
        operator_phones=["9111111111", "9222222222", "9333333333"],
        clones_per_operator=3,
    )
    assert len(chunks) == 9
    # Chunks 1-3 -> op1, 4-6 -> op2, 7-9 -> op3
    assert [c.operator_slot for c in chunks] == [
        "op1", "op1", "op1", "op2", "op2", "op2", "op3", "op3", "op3",
    ]
    assert [c.operator_phone for c in chunks] == [
        "9111111111", "9111111111", "9111111111",
        "9222222222", "9222222222", "9222222222",
        "9333333333", "9333333333", "9333333333",
    ]
    # Contiguous row buckets
    assert [c.row_count for c in chunks] == [3, 3, 3, 3, 3, 3, 3, 3, 3]


def test_split_multi_operator_uneven_rows(split_env):
    """N = 10 rows, K=3 M=3 -> 9 chunks. Effective size = ceil(10/9) = 2.
    Chunks are not strictly balanced; the implementation uses existing
    equal-contiguous-split logic (trailing chunks smaller or empty)."""
    inp = _make_input_xlsx(split_env / "Input" / "file.xlsx", n_rows=10)
    chunks = splitter.split(
        "LALJI", inp,
        operator_phones=["9111111111", "9222222222", "9333333333"],
        clones_per_operator=3,
    )
    assert len(chunks) == 9
    assert sum(c.row_count for c in chunks) == 10


def test_split_multi_operator_M_too_high_rejected(split_env):
    inp = _make_input_xlsx(split_env / "Input" / "file.xlsx", n_rows=20)
    with pytest.raises(ValueError, match="clones_per_operator"):
        splitter.split(
            "T", inp,
            operator_phones=["9111111111"],
            clones_per_operator=4,
        )


def test_split_multi_operator_M_zero_rejected(split_env):
    inp = _make_input_xlsx(split_env / "Input" / "file.xlsx", n_rows=20)
    with pytest.raises(ValueError, match="clones_per_operator"):
        splitter.split(
            "T", inp,
            operator_phones=["9111111111"],
            clones_per_operator=0,
        )


def test_split_legacy_single_operator_still_works(split_env):
    """Without operator_phones, split() behaves exactly as before:
    all chunks get slot=op1, phone=''."""
    inp = _make_input_xlsx(split_env / "Input" / "file.xlsx", n_rows=20)
    chunks = splitter.split("TEST", inp, chunk_size=5)
    assert len(chunks) == 4
    assert all(c.operator_slot == "op1" for c in chunks)
    assert all(c.operator_phone == "" for c in chunks)
